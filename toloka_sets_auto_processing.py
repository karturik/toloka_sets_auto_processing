import os
import pandas as pd
import requests
from pathlib import Path
import concurrent.futures
import random
import time
import datetime
from typing_extensions import Self

import toloka.client as toloka
import openpyxl
from countryinfo import CountryInfo
from geopy.geocoders import Nominatim
from functools import partial
import psycopg2
from colorama import Fore, Back, Style


validator_name = 'Validator name'

geolocator = Nominatim(user_agent="geoapiExercises")
geocode = partial(geolocator.geocode, language="en")

e = datetime.datetime.now()
date = f"{'%s-%s-%s' % (e.year, e.month, e.day)}"

conn = psycopg2.connect("""
    host=host
    port=port
    sslmode=require
    dbname=dbname
    user=user
    password=password
    target_session_attrs=read-write
""")
q = conn.cursor()

q.execute('''SELECT assignment_id, worker_id, assignment_nation, assignment_month,
                                              assignment_send_date, assignment_toloka_date, toloka_status,
                                              reward, account, pool_type, decision, reject_reason,
                                              hashes, gender FROM public.sets ''')
all_sets_in_db_df = pd.DataFrame(q.fetchall(), columns = ['assignment_id', 'worker_id', 'assignment_nation',
                                                          'assignment_month', 'assignment_send_date',
                                                          'assignment_toloka_date', 'toloka_status',
                                                          'reward', 'account', 'pool_type',
                                                          'decision', 'reject_reason', 'hashes', 'gender'])

URL_WORKER = 'https://toloka.yandex.ru/requester/worker/'
URL_API = "https://toloka.yandex.ru/api/v1/"
OAUTH_TOKEN = ''
HEADERS = {"Authorization": "OAuth %s" % OAUTH_TOKEN, "Content-Type": "application/JSON"}
toloka_client = toloka.TolokaClient(OAUTH_TOKEN, 'PRODUCTION')

account = ''
skill_id_reject = 1
skill_id_accept = 2

working_excel = pd.read_excel('sets_for_processing.xlsx', sheet_name='Лист1')

pool_number1 = 0

with open('need_manual.tsv', 'w', encoding='utf-8') as file:
    file.write(f"assignment_id\tempty field\n")
    file.close()

with open('errors.tsv', 'w', encoding='utf-8') as file:
    file.write(f"assignment_id\terror\n")
    file.close()

# START SET PROCESSING
for assignment_link in working_excel['assignment_link'].dropna():
    if '--' in assignment_link:
        tries = 0
        while tries < 10:
            # CAN BE FULL LINK, OR SHORT ASSIGNMENT_ID
            try:
                if not 'https://' in assignment_link and '--' in assignment_link:
                    assignment_id = assignment_link
                    pool_number = toloka_client.get_assignment(assignment_id=assignment_id).pool_id
                    project_id = toloka_client.get_pool(pool_id=pool_number).project_id
                    assignment_link = f'https://platform.toloka.ai/requester/project/{project_id}/pool/{pool_number}/assignments/{assignment_id}?direction=ASC'
                    print(assignment_link)
                    project_number = assignment_link.split('project/')[1].split('/pool')[0]
                    pool_number = assignment_link.split('/pool/')[1].split('/assignments')[0]
                    assignment_id = assignment_link.split('assignments/')[1].split('?direction')[0]
                    assignment_link = assignment_id
                else:
                    print(assignment_link)
                    project_number = assignment_link.split('project/')[1].split('/pool')[0]
                    pool_number = assignment_link.split('/pool/')[1].split('/assignments')[0]
                    assignment_id = assignment_link.split('assignments/')[1].split('?direction')[0]
                pool_name = toloka_client.get_pool(pool_id=pool_number).private_name
                # CLASSIFICATION POOL TYPE
                if 'new' in pool_name.lower() and not 'retry' in pool_name.lower() and not 'родствен' in pool_name.lower():
                    pool_type = 'new'
                elif 'retry' in pool_name.lower():
                    pool_type = 'retry'
                elif 'родствен' in pool_name.lower():
                    pool_type = 'родственники'
                else:
                    pool_type = ''
                print('project_number: ', project_number)
                print('pool_number: ', pool_number)
                print('assignment_id: ', assignment_id)
                assignment_request = toloka_client.get_assignment(assignment_id=assignment_id)
                if pool_number != pool_number1:
                    df_toloka = toloka_client.get_assignments_df(pool_number, status = ['APPROVED', 'SUBMITTED', 'REJECTED'])
                    # print(df_toloka.head(5))
                    pool_number1 = pool_number
                else:
                    df_toloka = df_toloka
                # CHECK IF WORKER HAS RACE IN OUTPUT
                if 'OUTPUT:race' in df_toloka:
                    try:
                        ethnicity = df_toloka[df_toloka['ASSIGNMENT:assignment_id'] == assignment_id]['OUTPUT:race'].values[0]
                        if ethnicity == 'Middle Eastern':
                            ethnicity = 'Middle East'
                        elif ethnicity == 'South Asian':
                            ethnicity = 'South Asia'
                        # nation_for_data_base = ethnicity
                    except Exception as e:
                        with open('need_manual.tsv', 'a', encoding='utf-8') as file:
                            file.write(f"{assignment_id}\tethnicity\n")
                            file.close()
                        with open('errors.tsv', 'a', encoding='utf-8') as file:
                            file.write(f"{assignment_id}\t{e}\n")
                            file.close()
                        ethnicity = ""
                else:
                    with open('need_manual.tsv', 'a', encoding='utf-8') as file:
                        file.write(f"{assignment_id}\tнациональность\n")
                        file.close()
                    ethnicity = ""
                print('ethnicity: ', ethnicity)
                # GET WORKER DATA AND LANGUAGE
                worker_id = df_toloka[df_toloka['ASSIGNMENT:assignment_id'] == assignment_id]['ASSIGNMENT:worker_id'].values[0]
                worker = requests.get(url='https://toloka.yandex.ru/api/new/requester/workers/' + worker_id, headers=HEADERS).json()
                print('worker_id: ', worker_id)

                if project_number == '123538' or project_number == '120426' or project_number == '120106':
                    if 'ES' in worker['languages']:
                        worker_language = 'ES'
                    else:
                        worker_language = 'EN'
                elif project_number == '123536' or project_number == '115606':
                    if 'RU' in worker['languages']:
                        worker_language = 'RU'
                    else:
                        worker_language = 'EN'
                elif project_number == '123537' or project_number == '123072' or project_number == '105897':
                    if 'ES' in worker['languages']:
                        worker_language = 'ES'
                    elif 'FR' in worker['languages']:
                        worker_language = 'FR'
                    elif 'EN' in worker['languages']:
                        worker_language = 'EN'
                    elif 'AR' in worker['languages']:
                        worker_language = 'AR'
                    elif 'ID' in worker['languages']:
                        worker_language = 'ID'
                    else:
                        worker_language = 'EN'
                else:
                    worker_language = 'EN'

                if 'OUTPUT:language' in df_toloka:
                    input_worker_language = df_toloka[df_toloka['ASSIGNMENT:assignment_id'] == assignment_id]['OUTPUT:language'].values[0]
                    if type(input_worker_language) != float:
                        if 'es' in input_worker_language:
                            worker_language = 'ES'
                        elif 'ru' in input_worker_language:
                            worker_language = 'RU'
                        elif 'fr' in input_worker_language:
                            worker_language = 'FR'
                        elif 'tr' in input_worker_language:
                            worker_language = 'TR'
                        elif 'ar' in input_worker_language:
                            worker_language = 'AR'
                        elif 'id' in input_worker_language:
                            worker_language = 'ID'
                else:
                    pass

                # SPECIFYING WORKER TEXTS, GET IT FROM WORK-EXCEL
                print('worker language: ', worker_language)
                if worker_language == 'RU':
                    message_text = working_excel['message'][0]
                    comment = 'Выполнено не по инструкции'
                    topic_reject = "Сделайте 10 фото лица"
                    topic_accept = "Большое вам спасибо за правильное выполнение задания, мы действительно ценим ваши усилия!"

                elif worker_language == 'ES':
                    message_text = working_excel['message_espanien'][0]
                    comment = 'No cumplida según las instrucciones'
                    topic_reject = "Toma 10 fotos faciales"
                    topic_accept = "Muchas gracias por completar la tarea correctamente, realmente apreciamos sus esfuerzos!"

                elif worker_language == 'FR':
                    message_text = working_excel['message_francusien'][0]
                    comment = 'Non rempli selon les instructions'
                    topic_reject = "Faire 10 photos de visage"
                    topic_accept = "Merci beaucoup pour la bonne exécution de la tâche, nous apprécions vraiment vos efforts!"

                elif worker_language == 'TR':
                    message_text = working_excel['message_turkey'][0]
                    comment = 'Talimatlara göre yapılmadı'
                    topic_reject = "Yüzün 10 fotoğrafını çekin"
                    topic_accept = "Görevi doğru tamamladığınız için çok teşekkür ederim, çabalarınızı gerçekten takdir ediyoruz!"

                elif worker_language == 'AR':
                    message_text = working_excel['message_arab'][0]
                    comment = 'لا تملأ وفقا للتعليمات'
                    topic_reject = "اصنع 10 صور للوجه"
                    topic_accept = "شكرا جزيلا لك على التنفيذ الجيد للمهمة ، ونحن نقدر حقا جهودك!"

                elif worker_language == 'ID':
                    message_text = working_excel['message_indonesian'][0]
                    comment = 'Tidak dilakukan sesuai dengan instruksi'
                    topic_reject = "Ambil 10 foto wajah"
                    topic_accept = "Terima kasih banyak telah menyelesaikan tugas dengan benar, kami sangat menghargai upaya Anda!"

                else:
                    message_text = working_excel['message_english'][0]
                    comment = 'Not according to the instructions'
                    topic_reject = "Take 10 photos of your face"
                    topic_accept = "Thank you very much for completing the task correctly, we really appreciate your efforts!"

                # GET REFUSAL REASONS FROM WORK-EXCEL
                li_refusal_reason_for_message = message_text.split('<ol>')[1].split('</ol>')[0]
                refusal_reason_for_message = ""
                refusal_reason_text_list = []
                refusal_reasons_number_list = working_excel[working_excel['assignment_link']==assignment_link]['refusal_reasons'].values[0]
                print(refusal_reasons_number_list)
                if "f" in str(refusal_reasons_number_list).lower():
                    sex = "FEMALE"
                elif "m" in str(refusal_reasons_number_list).lower():
                    sex = "MALE"
                else:sex = False
                refusal_reasons_number_list = str(refusal_reasons_number_list).replace("f", "").replace("m", "").replace(" f ", " ").replace(" m ", " ")

                if not "+" in refusal_reasons_number_list and not "-" in refusal_reasons_number_list and not "$" in refusal_reasons_number_list:

                    # SET REJECTION
                    if " " in refusal_reasons_number_list.strip():
                        refusal_reasons_number_list = refusal_reasons_number_list.replace("  ", " ").strip().split(" ")
                    else:
                        refusal_reasons_number_list = [int(float(refusal_reasons_number_list.strip()))]
                    print('refusal_reasons_list: ', refusal_reasons_number_list)
                    # GET ALL REJECT REASONS FROM EXCEL
                    for refusal_reason_number in refusal_reasons_number_list:
                        if worker_language == 'RU':
                            refusal_reason = working_excel[working_excel['refusal_reasons_number'] == int(refusal_reason_number)]['refusal_reasons_text'].values[0]
                        elif worker_language == 'ES':
                            refusal_reason = working_excel[working_excel['refusal_reasons_number'] == int(refusal_reason_number)]['refusal_reasons_text_espanien'].values[0]
                        elif worker_language == 'FR':
                            refusal_reason = working_excel[working_excel['refusal_reasons_number'] == int(refusal_reason_number)]['refusal_reasons_text_francusien'].values[0]
                        elif worker_language == 'TR':
                            refusal_reason = working_excel[working_excel['refusal_reasons_number'] == int(refusal_reason_number)]['refusal_reasons_text_turkish'].values[0]
                        elif worker_language == 'AR':
                            refusal_reason = working_excel[working_excel['refusal_reasons_number'] == int(refusal_reason_number)]['refusal_reasons_text_arab'].values[0]
                        elif worker_language == 'ID':
                            refusal_reason = working_excel[working_excel['refusal_reasons_number'] == int(refusal_reason_number)]['refusal_reasons_text_indonesian'].values[0]
                        else:
                            refusal_reason = working_excel[working_excel['refusal_reasons_number'] == int(refusal_reason_number)]['refusal_reasons_text_english'].values[0]
                        refusal_reason_text_list.append(refusal_reason)
                        refusal_reason_for_message = refusal_reason_for_message + li_refusal_reason_for_message.replace("#141825;'>", f"#141825;'> {refusal_reason}")
                    message_text = message_text.replace(li_refusal_reason_for_message, refusal_reason_for_message).replace('{ACCOUNT}', account).replace('{project_number}', project_number)
                    if df_toloka[df_toloka['ASSIGNMENT:assignment_id'] == assignment_id]['ASSIGNMENT:status'].values[0] != 'REJECTED' \
                            and df_toloka[df_toloka['ASSIGNMENT:assignment_id'] == assignment_id]['ASSIGNMENT:status'].values[0] != 'APPROVED':
                        toloka_client.reject_assignment(assignment_id=assignment_id, public_comment=comment)
                        print('Reject set')
                        # REJECT MESSAGE SENDING
                        message_body = {
                                      "topic": {
                                        "EN": topic_reject,
                                        "RU": topic_reject
                                      },
                                      "text": {
                                        "EN": message_text,
                                        "RU": message_text
                                      },
                                      "recipients_select_type": "DIRECT",
                                      "recipients_ids": [worker_id],
                                      "answerable": True
                                      }

                        requests.post(url='https://toloka.dev/api/v1/message-threads/compose', headers=HEADERS, data=message_body)

                        url = 'https://toloka.yandex.ru/api/v1/message-threads/compose'
                        send_msg = requests.post(url, headers=HEADERS, json=message_body).json()
                        if 'created' in send_msg:
                            print('Message sended')
                        else:
                            print('Message not send: ', send_msg)
                            with open('need_manual.tsv', 'a', encoding='utf-8') as file:
                                file.write(f"{assignment_id}\tотправить сообщение\n")
                                file.close()
                        # SKILL ADD
                        if not 404 in refusal_reasons_number_list and not '404' in refusal_reasons_number_list:
                            skill_body = {
                                "skill_id": skill_id_reject,
                                "user_id": worker_id,
                                "value": 60,
                                "reason": "Rehab 10 photos"
                            }
                            url = 'https://toloka.dev/api/v1/user-skills'
                            add_skill = requests.put(url, headers=HEADERS, json=skill_body).json()
                            if 'created' in add_skill:
                                print('Skill gived')
                            else:
                                print('No skill: ', add_skill)
                                with open('need_manual.tsv', 'a', encoding='utf-8') as file:
                                    file.write(f"{assignment_id}\tвыдать навык\n")
                                    file.close()

                    else:
                        print('Set already rejected: ', df_toloka[df_toloka['ASSIGNMENT:assignment_id'] == assignment_id]['ASSIGNMENT:status'].values[0])
                    if assignment_id in all_sets_in_db_df['assignment_id'].unique():
                        q.execute(f"UPDATE public.sets SET validator_name = '{validator_name}', validation_date = '{date}', decision = 'REJECTED', "
                                  f"toloka_status = 'REJECTED', reject_reason = '{';'.join(refusal_reason_text_list)}' WHERE assignment_id = '{assignment_id}';")
                        conn.commit()
                    print("-----------------------------------------------")
                    pass

                elif "-" in refusal_reasons_number_list or "-" in refusal_reasons_number_list:
                # REJECT SET WITHOUT GIVING RETRY SKILL TO USER
                    if df_toloka[df_toloka['ASSIGNMENT:assignment_id'] == assignment_id]['ASSIGNMENT:status'].values[0] != 'REJECTED':
                        toloka_client.reject_assignment(assignment_id=assignment_id, public_comment=comment)
                        print('Reject set')
                    else:
                        print('Set already rejected: ', df_toloka[df_toloka['ASSIGNMENT:assignment_id'] == assignment_id]['ASSIGNMENT:status'].values[0])
                    print("-----------------------------------------------")
                    pass

                elif refusal_reasons_number_list == "+" or "+" in refusal_reasons_number_list:
                    # DOWNLOAD SET
                    download_decision = True
                    # CHECK IF SAME SET IS IN DB
                    if assignment_id in all_sets_in_db_df['assignment_id'].unique():
                        print(Fore.RED + 'There is same set in DB: ' + Style.RESET_ALL)
                        assignment_link = f'https://platform.toloka.ai/requester/project/{project_number}/pool/{pool_number}/assignments/{assignment_id}?direction=ASC'
                        print(' - link to set: ', assignment_link)
                        
                        print(Fore.RED + f"    ({assignment_id} {Style.RESET_ALL}, {all_sets_in_db_df[all_sets_in_db_df['assignment_id']==assignment_id]['worker_id'].values[0]},"
                                         f" {all_sets_in_db_df[all_sets_in_db_df['assignment_id']==assignment_id]['assignment_nation'].values[0]},"
                                         f" {all_sets_in_db_df[all_sets_in_db_df['assignment_id']==assignment_id]['assignment_month'].values[0]},"
                                         f" {all_sets_in_db_df[all_sets_in_db_df['assignment_id']==assignment_id]['assignment_send_date'].values[0]},"
                                         f" {all_sets_in_db_df[all_sets_in_db_df['assignment_id']==assignment_id]['assignment_toloka_date'].values[0]},"
                                         f"{Fore.RED} toloka-status: {all_sets_in_db_df[all_sets_in_db_df['assignment_id']==assignment_id]['toloka_status'].values[0]}{Style.RESET_ALL},"
                                         f" {all_sets_in_db_df[all_sets_in_db_df['assignment_id']==assignment_id]['pool_type'].values[0]},"
                                         f" {Fore.RED} status: {all_sets_in_db_df[all_sets_in_db_df['assignment_id']==assignment_id]['decision'].values[0]}"
                                         f" {Style.RESET_ALL}, {all_sets_in_db_df[all_sets_in_db_df['assignment_id']==assignment_id]['reject_reason'].values[0]})" + Style.RESET_ALL)

                        print(Fore.GREEN + 'There are same set, with same assignment_id' + Style.RESET_ALL)
                        download_decision = input('Download? \n1.Yes \n2.No, skip set \n')
                        if download_decision == '1':
                            download_decision = True
                        else:
                            download_decision = False
                    # IF THAT IS UNIQUE SET => DOWNLOAD
                    if not assignment_id in all_sets_in_db_df['assignment_id'].unique() or download_decision == True:
                        download_decision = False
                        session = requests.Session()
                        session.headers.update(
                            {"Authorization": "OAuth %s" % OAUTH_TOKEN, "Content-Type": "application/JSON"}
                        )
                        class User():
                            def __init__(self: Self,
                                         workerId: str,
                                         age: int,
                                         country: str,
                                         gender: str = None,
                                         cityId: int = None) -> None:
                                self.workerId = workerId
                                self.age = age
                                self.country = self._translate_country(country)
                                self.gender = gender
                                self.city = self._translate_city(cityId)

                            def _translate_city(self, cityId, session=session):
                                if cityId is None:
                                    return None
                                else:
                                    r = session.get(f"https://toloka.yandex.ru/api/ctx/geobase/regions/{cityId}")
                                    city = r.json()["name"]
                                    return city

                            def _translate_country(self, country: str, session=session):
                                if country is None:
                                    return None
                                else:
                                    r = session.get("https://yastatic.net/s3/toloka/p/requester/toloka.en.f8134e9b10b398400049.json")
                                    country = r.json()[f"country:{country}"]
                                    return country

                            def __repr__(self) -> str:
                                age = self.age
                                country = self.country
                                city = self.city
                                gender = self.gender
                                return f"User({age=}, {country=}, {city=}, {gender=})"

                            def __str__(self) -> str:
                                age = self.age
                                country = self.country
                                city = self.city
                                gender = self.gender
                                return f'User complete'

                        def generate_metadata(user: User, path_to_file: Path):
                            columns = ['ID', 'Ethnicity', "Capturing Source", "Shooting location", "Document type", "Gender",
                                       "Age at date captured", "Location_country", "Location_state", "City", "State", "Country",
                                       "City", "State", "Country",
                                       "Filename"]
                            tmp = [None] * 10
                            df = [
                                [1, None, "Mobile", "Indoor", "National_Identity_Card"] + tmp + ["National_Identity_Card_.jpg"],
                                [2, None, "Mobile", "Indoor", "Passport"] + tmp + ["Passport_.jpg"],
                                [3, None, "Mobile", "Indoor", "None"] + tmp + ["Phone_indoor_01.jpg"],
                                [4, None, "Mobile", "Indoor", "None"] + tmp + ["Phone_indoor_02.jpg"],
                                [5, None, "Mobile", "Outdoor", "None"] + tmp + ["Phone_outdoor_01.jpg"],
                                [6, None, "Mobile", "Outdoor", "None"] + tmp + ["Phone_outdoor_02.jpg"],
                                [7, None, "Webcam", "Indoor", "None"] + tmp + ["Webcam_01.jpg"],
                                [8, None, "Webcam", "Indoor", "None"] + tmp + ["Webcam_02.jpg"],
                                [9, None, "Webcam", "Indoor", "None"] + tmp + ["Webcam_03.jpg"],
                                [10, None, "Webcam", "Indoor", "None"] + tmp + ["Webcam_04.jpg"],
                            ]
                            if user.age > 18:
                                passport_age = user.age - random.randint(0, (user.age - 18))
                                card_age = passport_age + random.randint(0, (user.age - passport_age))
                            else:
                                passport_age = user.age
                                card_age = user.age
                            # GET COUNTRY FROM LIB FOR AUTOMATION METADATA FILLING
                            if len(user.country) > 2:
                                try:
                                    if not user.city or len(user.city) < 2:
                                        capital = CountryInfo(user.country).info()['capital']
                                        city = capital
                                    else:
                                        city = user.city
                                    state = geocode(city).address
                                except:
                                    try:
                                        if not user.city or len(user.city) < 2:
                                            capital = CountryInfo(user.country).info()['capital']
                                            city = capital
                                        else:
                                            city = user.city
                                        state = geocode(city).address.encode(encoding = 'UTF-8', errors = 'strict')
                                    except Exception as e:
                                        with open('errors.tsv', 'a', encoding='utf-8') as file:
                                            file.write(f"{assignment_id}\t{e}\n")
                                            file.close()
                                        city = ""
                                        state = ""
                                        with open('need_manual.tsv', 'a', encoding='utf-8') as file:
                                            file.write(f"{assignment_id}\tгород и штат\n")
                                            file.close()
                            else:
                                user.country = ""
                                city = ""
                                state = ""
                                with open('need_manual.tsv', 'a', encoding='utf-8') as file:
                                    file.write(f"{assignment_id}\tстрану, город и штат\n")
                                    file.close()

                            if not user.gender:
                                if sex:
                                    user.gender = sex
                                else:
                                    with open('need_manual.tsv', 'a', encoding='utf-8') as file:
                                        file.write(f"{assignment_id}\tгендер\n")
                                        file.close()
                            else:
                                if len(user.gender) < 2:
                                    if sex:
                                        user.gender = sex
                                    else:
                                        with open('need_manual.tsv', 'a', encoding='utf-8') as file:
                                            file.write(f"{assignment_id}\tгендер\n")
                                            file.close()

                            df = pd.DataFrame(df, columns=columns)
                            df.loc[:, "Gender"] = user.gender
                            df.loc[0, "Age at date captured"] = passport_age
                            df.loc[1, "Age at date captured"] = card_age
                            df.loc[2:, "Age at date captured"] = user.age
                            df.loc[:, "Location_country"] = user.country
                            df.loc[:, "Location_state"] = state
                            df.loc[:, "State"] = state
                            df.loc[:, "City"] = city
                            df.loc[:, "Country"] = user.country
                            df.loc[:, "Ethnicity"] = ethnicity
                            df.to_excel(f"{path_to_file}/metadata.xlsx", startrow=1, index=False)

                            # MERGING CELLS
                            wb = openpyxl.load_workbook(f"{path_to_file}/metadata.xlsx")
                            sheet = wb.active

                            sheet.cell(column=8, row=1, value="ID details")
                            sheet.cell(column=11, row=1, value="Place of capture")
                            sheet.cell(column=13, row=1, value="Place of birth")

                            sheet.merge_cells("H1:J1")
                            sheet.merge_cells("K1:L1")
                            sheet.merge_cells("M1:O1")
                            wb.save(f"{path_to_file}/metadata.xlsx")

                        IMAGE_NAMES = [
                            "National_Identity_Card_.jpg",
                            "Passport_.jpg",
                            "Phone_indoor_01.jpg",
                            "Phone_indoor_02.jpg",
                            "Phone_outdoor_01.jpg",
                            "Phone_outdoor_02.jpg",
                            "Webcam_01.jpg",
                            "Webcam_02.jpg",
                            "Webcam_03.jpg",
                            "Webcam_04.jpg",
                        ]

                        df_toloka_one_row = df_toloka[df_toloka['ASSIGNMENT:assignment_id'] == assignment_id]
                        try:
                            del df_toloka_one_row['OUTPUT:race']
                            del df_toloka_one_row['GOLDEN:race']
                        except:
                            pass
                        print("Getting data...")
                        files_rows = list(map(list, [i for i in df_toloka_one_row.loc[:, ['OUTPUT:img_1','OUTPUT:img_2','OUTPUT:img_3',
                                                                                          'OUTPUT:img_4','OUTPUT:img_5','OUTPUT:img_6',
                                                                                          'OUTPUT:img_7','OUTPUT:img_8','OUTPUT:img_9',
                                                                                          'OUTPUT:img_10']].values]))
                        assignment_ids = [i for i in df_toloka_one_row.loc[:, ['ASSIGNMENT:assignment_id']].values]
                        users_ids = [i for i in df_toloka_one_row.loc[:, ['ASSIGNMENT:worker_id']].values]
                        print("-----------------------------------------------")

                        user_count = 0
                        def user_image_download(index, image_id):
                            data = session.get(
                                f"https://toloka.dev/api/v1/attachments/{image_id}/download"
                            )
                            with open(path_to_images / IMAGE_NAMES[index], "wb") as f:
                                f.write(data.content)
                                return (f"Finished {index + 1}-е/{len(file_row)} image.")

                        for assignment_id1, file_row, user_id in zip(assignment_ids, files_rows, users_ids):
                            if not os.path.exists('new_sets'):
                                os.mkdir('new_sets')
                            path_to_images = Path(f'new_sets/{assignment_id}').resolve()
                            path_to_images.mkdir(exist_ok=True)
                            print(f"Created dir '{assignment_id}'.")
                            print("Download images...")

                            with concurrent.futures.ThreadPoolExecutor() as executor:
                                futures = []
                                for index, image_id in enumerate(file_row):
                                    futures.append(executor.submit(user_image_download, index=index, image_id=image_id))
                                for future in concurrent.futures.as_completed(futures):
                                    print(future.result())

                            user_count += 1
                            print(f"All images finished {user_count}/{len(files_rows)} from user. Metadata generation...")
                            print(f"User: {worker_id}")
                            # r = session.get(f"https://toloka.yandex.ru/api/new/requester/workers/{user_id}")
                            user = User(
                                worker.get("workerId"),
                                worker.get("age"),
                                worker.get("country"),
                                worker.get("gender"),
                                worker.get("cityId")
                            )
                            print(user)
                            generate_metadata(user, path_to_images)
                            print(f"User(age={user.age}, country={user.country}, sex={user.gender}, ethnicity={ethnicity})")
                    print(f"Metadata generated! User finished")
                    print("-----------------------------------------------")

                tries = 10
            except Exception as e:
                if 'DoesNotExistApiError' in str(e):
                    # SCRIPT WORKING ON TWO ACCOUNTS
                    print('Change account')
                    if OAUTH_TOKEN == '':
                        OAUTH_TOKEN = ''
                        account = ''
                        skill_id_reject = 1
                        skill_id_accept = 2
                    elif OAUTH_TOKEN == '':
                        OAUTH_TOKEN = ''
                        account = ''
                        skill_id_reject = 3
                        skill_id_accept = 4
                    HEADERS = {"Authorization": "OAuth %s" % OAUTH_TOKEN, "Content-Type": "application/JSON"}
                    toloka_client = toloka.TolokaClient(OAUTH_TOKEN, 'PRODUCTION')
                tries += 1
                print(f'Error, try {tries}/10')
                time.sleep(1)
                if tries == 10:
                    with open('errors.tsv', 'a', encoding='utf-8') as file:
                        file.write(f"{assignment_id}\t{e}\n")
                        file.close()
    else:
        print(assignment_link)
        print('No link or assignment_id detected')
        print("-----------------------------------------------")

conn.close()

if os.path.exists('need_manual.tsv'):
    manual_df = pd.read_csv('need_manual.tsv', sep='\t')
    manual_count = len(manual_df['assignment_id'])
else:
    manual_count = 0

if os.path.exists('errors.tsv'):
    error_df = pd.read_csv('errors.tsv', sep='\t')
    error_count = len(error_df['assignment_id'])
else:
    error_count = 0

print(f'Need to manual filling: {manual_count}, file: need_manual.tsv')
print(f'Error count: {error_count}, file: errors.tsv')